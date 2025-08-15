"""File processing utilities for extracting text from various file formats."""

import io
import os
from typing import Tuple, Optional
from pathlib import Path

import PyPDF2
import fitz  # PyMuPDF
from docx import Document
from openpyxl import load_workbook
from PIL import Image
import pytesseract

from app.config import settings


class FileProcessor:
    """Handles text extraction from various file formats."""
    
    @staticmethod
    def detect_file_type(filename: str, content: bytes) -> str:
        """Detect file type from filename and content."""
        extension = Path(filename).suffix.lower()
        
        # Map extensions to file types
        type_mapping = {
            '.txt': 'text',
            '.pdf': 'pdf',
            '.docx': 'docx',
            '.doc': 'docx',  # Treat as docx for simplicity
            '.xlsx': 'xlsx',
            '.xls': 'xlsx',  # Treat as xlsx for simplicity
            '.png': 'image',
            '.jpg': 'image',
            '.jpeg': 'image',
            '.gif': 'image',
            '.bmp': 'image',
            '.tiff': 'image',
        }
        
        return type_mapping.get(extension, 'unknown')
    
    @staticmethod
    def validate_file(filename: str, content: bytes) -> Tuple[bool, Optional[str]]:
        """Validate file type and size."""
        # Check file size
        if len(content) > settings.max_file_size_bytes:
            return False, f"File size exceeds {settings.max_file_size_mb}MB limit"
        
        # Check file type
        file_type = FileProcessor.detect_file_type(filename, content)
        if file_type == 'unknown':
            extension = Path(filename).suffix.lower()
            return False, f"Unsupported file type: {extension}"
        
        # Check if file type is allowed
        extension = Path(filename).suffix.lower().lstrip('.')
        if extension not in settings.allowed_file_types_list:
            return False, f"File type '{extension}' not allowed. Allowed types: {settings.allowed_file_types}"
        
        return True, None
    
    @staticmethod
    def extract_text_from_pdf(content: bytes) -> str:
        """Extract text from PDF using PyMuPDF (fallback to PyPDF2)."""
        try:
            # Try PyMuPDF first (better text extraction)
            doc = fitz.open(stream=content, filetype="pdf")
            text = ""
            for page in doc:
                text += page.get_text()
            doc.close()
            return text.strip()
        except Exception:
            # Fallback to PyPDF2
            try:
                pdf_reader = PyPDF2.PdfReader(io.BytesIO(content))
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text()
                return text.strip()
            except Exception as e:
                raise ValueError(f"Failed to extract text from PDF: {str(e)}")
    
    @staticmethod
    def extract_text_from_docx(content: bytes) -> str:
        """Extract text from DOCX file."""
        try:
            doc = Document(io.BytesIO(content))
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text.strip()
        except Exception as e:
            raise ValueError(f"Failed to extract text from DOCX: {str(e)}")
    
    @staticmethod
    def extract_text_from_xlsx(content: bytes) -> str:
        """Extract text from XLSX file."""
        try:
            workbook = load_workbook(io.BytesIO(content), data_only=True)
            text = ""
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
                text += "\n"
            
            return text.strip()
        except Exception as e:
            raise ValueError(f"Failed to extract text from XLSX: {str(e)}")
    
    @staticmethod
    def extract_text_from_image(content: bytes) -> str:
        """Extract text from image using OCR (Tesseract)."""
        try:
            image = Image.open(io.BytesIO(content))
            
            # Convert to RGB if necessary
            if image.mode != 'RGB':
                image = image.convert('RGB')
            
            # Use Tesseract OCR to extract text
            text = pytesseract.image_to_string(image)
            return text.strip()
        except Exception as e:
            raise ValueError(f"Failed to extract text from image using OCR: {str(e)}")
    
    @staticmethod
    def extract_text(filename: str, content: bytes) -> Tuple[str, str]:
        """Extract text from file based on its type."""
        file_type = FileProcessor.detect_file_type(filename, content)
        
        if file_type == 'text':
            try:
                text = content.decode('utf-8')
            except UnicodeDecodeError:
                try:
                    text = content.decode('latin-1')
                except UnicodeDecodeError:
                    raise ValueError("Unable to decode text file")
        
        elif file_type == 'pdf':
            text = FileProcessor.extract_text_from_pdf(content)
        
        elif file_type == 'docx':
            text = FileProcessor.extract_text_from_docx(content)
        
        elif file_type == 'xlsx':
            text = FileProcessor.extract_text_from_xlsx(content)
        
        elif file_type == 'image':
            text = FileProcessor.extract_text_from_image(content)
        
        else:
            raise ValueError(f"Unsupported file type: {file_type}")
        
        if not text.strip():
            raise ValueError(f"No text could be extracted from the {file_type} file")
        
        return text, file_type
